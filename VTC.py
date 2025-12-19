import requests
import sys
import time
import openpyxl # Replaces xlrd and xlwt

# --- CONFIGURATION ---
# PASTE YOUR API KEY INSIDE THE QUOTES BELOW
API_KEY = 'XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX' 
OUTPUT_FILENAME = 'HashConvertedOutput.xlsx'

if not API_KEY:
    print("Error: Please open the script and paste your VirusTotal API Key.")
    sys.exit()

try:
    input_file = sys.argv[1]
except IndexError:
    print("Usage: python3 VTC.py <input_file.xlsx>")
    sys.exit()

print(f"Opening {input_file}...")

# 1. Open Input Excel Workbook (using openpyxl)
try:
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
except Exception as e:
    print(f"Error opening file: {e}")
    sys.exit()

# 2. Setup New Workbook for output
wbwrite = openpyxl.Workbook()
sheet1 = wbwrite.active
sheet1.title = 'Hashes'

# Write Headers
headers = ['Input Hash', 'MD5', 'SHA-1', 'SHA-256', 'Detections', 'Total AVs', 'Status']
sheet1.append(headers)

# VT API URL
url = 'https://www.virustotal.com/vtapi/v2/file/report'

# 3. Iterate through rows
# values_only=True yields the cell value directly
row_count = 0
total_rows = sheet.max_row

for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
    current_hash = row[0]
    
    # Skip empty rows
    if not current_hash:
        continue

    # Clean the hash (remove whitespace)
    current_hash = str(current_hash).strip()
    
    params = {'apikey': API_KEY, 'resource': current_hash}
    
    try:
        response = requests.get(url, params=params)
        
        # Check if HTTP request was successful
        if response.status_code == 204:
            print("(!) API Request Limit Exceeded. Waiting longer...")
            time.sleep(60) # Wait a full minute if quota hit
            continue
            
        data = response.json()
        
        # VirusTotal Response Code: 1 means found, 0 means not found
        response_code = data.get("response_code")

        if response_code == 1:
            md5 = data.get("md5", "N/A")
            sha1 = data.get("sha1", "N/A")
            sha256 = data.get("sha256", "N/A")
            positives = data.get("positives", 0)
            total = data.get("total", 0)
            status = "Found"
            
            # Write to output row
            sheet1.append([current_hash, md5, sha1, sha256, positives, total, status])
            print(f"[{row_count+1}/{total_rows}] Found: {current_hash}")
            
        else:
            # Hash not found in VT database
            sheet1.append([current_hash, "-", "-", "-", "-", "-", "Not Found in VT"])
            print(f"[{row_count+1}/{total_rows}] Not Found: {current_hash}")

    except Exception as e:
        print(f"Error processing {current_hash}: {e}")
        sheet1.append([current_hash, "ERROR", "", "", "", "", str(e)])

    row_count += 1
    
    # Save periodically (optional, but good for long lists)
    if row_count % 10 == 0:
        wbwrite.save(OUTPUT_FILENAME)

    # VirusTotal Public API Sleep (15s buffer to be safe)
    time.sleep(16) 

# Final Save
wbwrite.save(OUTPUT_FILENAME)
print(f"\nCompleted. Data saved to {OUTPUT_FILENAME}")
